"""
modules/xlsx_builder.py
Multi-sheet XLSX QC workbook builder.

Exports build_xlsx(df, mapping, results, sas_logic, xlsx_path).
"""

import re

import pandas as pd

from .utils import col, col_list, is_blank


def _df_from_records(records):
    return pd.DataFrame(records) if records else pd.DataFrame()


def _write_sheet(wb, sheet_name, df, highlight_col=None, highlight_val="Y",
                 highlight_color="FFF2CC"):
    """Write a DataFrame to an openpyxl workbook sheet with optional row highlighting."""
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    ws          = wb.create_sheet(title=sheet_name[:31])
    header_fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    if df.empty:
        ws.append(["No records"])
        return

    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.fill = header_fill
                cell.font = header_font
        elif highlight_col and highlight_col in df.columns:
            col_idx = list(df.columns).index(highlight_col) + 1
            if str(ws.cell(r_idx, col_idx).value) == highlight_val:
                ff = PatternFill(start_color=highlight_color,
                                 end_color=highlight_color, fill_type="solid")
                for cell in ws[r_idx]:
                    cell.fill = ff

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)


def build_xlsx(df, mapping, results, sas_logic, xlsx_path):
    """Build the full multi-sheet QC workbook and save to xlsx_path."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    pksumxfl_crits = sas_logic["pksumxfl_crits"]
    ncaxfl_ncaxrs  = sas_logic["ncaxfl_ncaxrs"]

    uid    = col(df, "USUBJID",  mapping)
    test   = col(df, "PCTEST",   mapping)
    visit  = col(df, "VISITCD",  mapping) or col(df, "AVISIT", mapping)
    nca01  = col(df, "NCA01FL",  mapping)
    pksumx = col(df, "PKSUMXFL", mapping)
    mrrlt  = col(df, "MRRLT",    mapping)
    nrrlt  = col(df, "NRRLT",    mapping)
    aval   = col(df, "AVAL",     mapping)
    dtype  = col(df, "DTYPE",    mapping)
    aresc  = col(df, "ARESC",    mapping)

    crit_cols   = col_list(df, pksumxfl_crits)
    all_crit_fl = [c for c in df.columns if re.match(r"^CRIT\d+FL$", c, re.I)]
    all_nca_xrs = [c for c in df.columns if re.match(r"^NCA\d+XRS$", c, re.I)]

    listing_cols = [c for c in [uid, test, visit, mrrlt, nrrlt, aval, aresc, nca01, pksumx, dtype] if c]
    listing_cols += [c for c in all_crit_fl if c not in listing_cols]
    listing_cols += [c for c in all_nca_xrs if c not in listing_cols]

    # ── Core sheets ───────────────────────────────────────────────────────────
    _write_sheet(wb, "CRIT_NCA_Crossref",
                 _df_from_records(results.get("crit_nca_crossref", {}).get("crossref", [])))

    logic_rows = [
        {"Derived_Flag": "PKSUMXFL = 'Y'", "SAS_Function": "whichc('Y', ...)",
         "Variables": ", ".join(pksumxfl_crits),
         "Explanation": "Record excluded from summary stats if ANY listed CRIT = 'Y'.",
         "Parse_Method": sas_logic["parse_method"]},
        {"Derived_Flag": "NCAXFL = 'Y'", "SAS_Function": "cmiss(...) < N",
         "Variables": ", ".join(ncaxfl_ncaxrs),
         "Explanation": "Record flagged if ANY listed NCAxxXRS is non-empty.",
         "Parse_Method": sas_logic["parse_method"]},
        {"Derived_Flag": "NCA01FL = 'Y'", "SAS_Function": "compound",
         "Variables": sas_logic["nca01fl_logic"],
         "Explanation": "Record included in NCA if RICHRFL='Y' AND NCAXFL=' '.",
         "Parse_Method": sas_logic["parse_method"]},
    ]
    _write_sheet(wb, "Flag_Derivation_Logic", _df_from_records(logic_rows))
    _write_sheet(wb, "Treatment_Mapping", _df_from_records(results.get("treatment_mapping", {}).get("table", [])))
    _write_sheet(wb, "Route_Table",       _df_from_records(results.get("route_table", {}).get("table", [])))

    if nca01:
        nca_df = df[df[nca01].astype(str).str.strip() == "Y"][listing_cols].copy()
        if crit_cols:
            nca_df["ANY_CRIT_Y"] = nca_df[crit_cols].apply(
                lambda r: "Y" if any(str(v).strip() == "Y" for v in r) else "", axis=1
            )
        _write_sheet(wb, "NCA01FL_Y", nca_df, highlight_col="ANY_CRIT_Y", highlight_color="FFF2CC")

    if nca01 and crit_cols:
        any_crit = df[crit_cols].apply(lambda r: any(str(v).strip() == "Y" for v in r), axis=1)
        mask     = (df[nca01].astype(str).str.strip() == "Y") & any_crit
        _write_sheet(wb, "NCA01FL_with_CRIT", df[mask][listing_cols].copy(),
                     highlight_col=crit_cols[0] if crit_cols else None)

    if pksumx and crit_cols:
        any_crit = df[crit_cols].apply(lambda r: any(str(v).strip() == "Y" for v in r), axis=1)
        mask     = is_blank(df[pksumx]) & any_crit
        _write_sheet(wb, "Summary_Pop_with_CRIT", df[mask][listing_cols].copy(), highlight_color="FFCCCC")

    if pksumx and nca01:
        mask_trough = is_blank(df[pksumx]) & (df[nca01].astype(str).str.strip() != "Y")
        _write_sheet(wb, "Trough_Records", df[mask_trough][listing_cols].copy())

    dup_res = results.get("mrrlt_duplicates", {})
    for pop_key, sheet_label in [("all_records", "MRRLT_Dups_All"),
                                  ("nca_population", "MRRLT_Dups_NCA"),
                                  ("summary_population", "MRRLT_Dups_Summary")]:
        _write_sheet(wb, sheet_label, _df_from_records(dup_res.get(pop_key, {}).get("records", [])),
                     highlight_color="FFD700")

    err_res = results.get("erroneous_records", {})
    all_err = []
    for key, val in err_res.items():
        if isinstance(val, dict) and val.get("records"):
            for row in val["records"]:
                row["_source_check"] = key
            all_err.extend(val["records"])
    _write_sheet(wb, "Erroneous_Records", _df_from_records(all_err), highlight_color="FFCCCC")

    if dtype:
        _write_sheet(wb, "DELETED_Records",
                     df[df[dtype].astype(str).str.strip() == "DELETED"][listing_cols].copy(),
                     highlight_color="FFC7CE")

    dose_flags = col_list(df, ["CRIT3FL", "MISSED_DOSE_FL", "CRIT7FL", "RECORD_INCOMPLETE_DOSE"])
    if dose_flags:
        any_dose = df[dose_flags].apply(lambda r: any(str(v).strip() == "Y" for v in r), axis=1)
        _write_sheet(wb, "Missing_Dose", df[any_dose][listing_cols].copy(), highlight_color="FCE4D6")

    vom_flags = col_list(df, ["CRIT12FL", "VOMFL"])
    if vom_flags:
        any_vom = df[vom_flags].apply(lambda r: any(str(v).strip() == "Y" for v in r), axis=1)
        _write_sheet(wb, "Vomiting", df[any_vom][listing_cols].copy(), highlight_color="E2EFDA")

    sc = results.get("subject_counts", {})
    for key, sheet in [("by_phase", "SubjectCounts_Phase"),
                        ("by_actarm", "SubjectCounts_ARM"),
                        ("by_pctest", "SubjectCounts_PCTEST")]:
        _write_sheet(wb, sheet, _df_from_records(sc.get(key, [])))

    pt = results.get("profile_types", {})
    _write_sheet(wb, "Profile_Types_Overall", _df_from_records(pt.get("overall_distribution", [])))
    if pt.get("by_group"):
        _write_sheet(wb, "Profile_Types_ByGroup", _df_from_records(pt["by_group"]))

    contra = results.get("populations", {}).get("population_contradiction", {})
    if contra.get("records"):
        _write_sheet(wb, "Pop_Contradiction", _df_from_records(contra["records"]), highlight_color="FF0000")

    specs = results.get("specs_comparison", {})
    if specs.get("specs_vars"):
        ds_cols_upper = {c.upper() for c in df.columns}
        specs_rows = [{"Name": v["Name"], "Label": v.get("Label", ""), "Core": v.get("Core", ""),
                       "In_Dataset": "YES" if v["Name"] in ds_cols_upper else "NO"}
                      for v in specs["specs_vars"]]
        _write_sheet(wb, "Specs_Comparison", _df_from_records(specs_rows),
                     highlight_col="In_Dataset", highlight_val="NO", highlight_color="FFCCCC")

    ap = results.get("aprofile_alignment", {})
    if ap.get("case1_records"):
        _write_sheet(wb, "APROFILE_NCA01_missing", _df_from_records(ap["case1_records"]), highlight_color="FF0000")
    if ap.get("case2_records"):
        _write_sheet(wb, "APROFILE_NCA01_mismatch", _df_from_records(ap["case2_records"]), highlight_color="FFF2CC")
    if ap.get("profile_summary"):
        _write_sheet(wb, "APROFILE_Summary", _df_from_records(ap["profile_summary"]))

    # ── Gap check sheets ──────────────────────────────────────────────────────
    def _gap(key, subkey=None, sheet=None, color="FFCCCC"):
        data = results.get(key, {})
        rows = data.get(subkey, []) if subkey else data.get("records", [])
        if rows:
            _write_sheet(wb, sheet or key[:31], _df_from_records(rows), highlight_color=color)

    _gap("phase3_exclusion",  sheet="Phase3_Records",    color="FF0000")
    _gap("astx030_exclusion", sheet="ASTX030_Records",   color="FF0000")
    _gap("sort_key_uniqueness", sheet="SortKey_Duplicates", color="FF0000")

    av = results.get("avalu_consistency", {})
    if av.get("avalu_distribution"):
        _write_sheet(wb, "AVALU_Distribution", _df_from_records(av["avalu_distribution"]))
    if av.get("records_aval_no_avalu"):
        _write_sheet(wb, "AVALU_Missing", _df_from_records(av["records_aval_no_avalu"]), highlight_color="FFCCCC")

    bc = results.get("bsacat_derivation", {})
    if bc.get("mismatch_records"):
        _write_sheet(wb, "BSACAT_Errors", _df_from_records(bc["mismatch_records"]), highlight_color="FFCCCC")

    ap2 = results.get("aperiodc_derivation", {})
    if ap2.get("mismatch_records"):
        _write_sheet(wb, "APERIODC_Errors", _df_from_records(ap2["mismatch_records"]), highlight_color="FFCCCC")

    aseq_r = results.get("aseq_derivation", {})
    if aseq_r.get("sequence_aseq_crosstab"):
        _write_sheet(wb, "ASEQ_Crosstab", _df_from_records(aseq_r["sequence_aseq_crosstab"]))
    if aseq_r.get("mismatch_records"):
        _write_sheet(wb, "ASEQ_Errors", _df_from_records(aseq_r["mismatch_records"]), highlight_color="FFCCCC")

    ccd = results.get("cohortcd_derivation", {})
    if ccd.get("cohort_cohortcd_crosstab"):
        _write_sheet(wb, "COHORTCD_Crosstab", _df_from_records(ccd["cohort_cohortcd_crosstab"]))

    dcat = results.get("drugcat_derivation", {})
    if dcat.get("phase_cohortcd_drugcat_crosstab"):
        _write_sheet(wb, "DRUGCAT_Crosstab", _df_from_records(dcat["phase_cohortcd_drugcat_crosstab"]))
    if dcat.get("phase2_drugcat_set_records"):
        _write_sheet(wb, "DRUGCAT_Ph2_Error", _df_from_records(dcat["phase2_drugcat_set_records"]), highlight_color="FFCCCC")

    ti = results.get("trtint_derivation", {})
    if ti.get("trtint_distribution"):
        _write_sheet(wb, "TRTINT_Distribution", _df_from_records(ti["trtint_distribution"]))
    if ti.get("records_should_be_24_but_not"):
        _write_sheet(wb, "TRTINT_Missing24", _df_from_records(ti["records_should_be_24_but_not"]), highlight_color="FFCCCC")
    if ti.get("records_unexpected_24"):
        _write_sheet(wb, "TRTINT_Unexpected24", _df_from_records(ti["records_unexpected_24"]), highlight_color="FFF2CC")

    n40 = results.get("nca40xrs_consistency", {})
    if n40.get("miss_fire_records"):
        _write_sheet(wb, "NCA40XRS_MissFire",  _df_from_records(n40["miss_fire_records"]),  highlight_color="FFCCCC")
    if n40.get("false_fire_records"):
        _write_sheet(wb, "NCA40XRS_FalseFire", _df_from_records(n40["false_fire_records"]), highlight_color="FFF2CC")

    cso = results.get("crit_subject_overrides", {})
    if isinstance(cso.get("crit2fl_overrides"), list):
        _write_sheet(wb, "CRIT2FL_Overrides", _df_from_records(cso["crit2fl_overrides"]))

    mp = results.get("manual_patches", {})
    if mp.get("patches"):
        patch_rows = [{"Patch": p.get("patch"), "Subject": p.get("subject"), "VISITCD": p.get("visitcd"),
                       "Description": p.get("description"), "N_records": p.get("n_records", 0),
                       "DOSEA_errors": p.get("n_dosea_not_20", "N/A"),
                       "MNRRLT_FL_errors": p.get("n_mnrrlt_fl_not_y", "N/A"),
                       "MRRLT_errors": p.get("n_mrrlt_ne_nrrlt", "N/A")}
                      for p in mp["patches"]]
        _write_sheet(wb, "Manual_Patches", _df_from_records(patch_rows))

    ws_r = results.get("wide_summaries", {})
    tp_r = ws_r.get("timepoints_per_profile", {})
    if isinstance(tp_r, dict) and tp_r.get("sparse_profiles"):
        _write_sheet(wb, "Sparse_Profiles", _df_from_records(tp_r["sparse_profiles"]), highlight_color="FFF2CC")
    sv_r = ws_r.get("subject_per_visit", [])
    if isinstance(sv_r, list) and sv_r:
        _write_sheet(wb, "Subjects_Per_Visit", _df_from_records(sv_r))

    aif = results.get("aval_increase_flags", {})
    for crit_lbl in ["CRIT16FL", "CRIT17FL"]:
        info = aif.get(crit_lbl, {})
        if info.get("increase_not_flagged"):
            _write_sheet(wb, f"{crit_lbl}_Missed",   _df_from_records(info["increase_not_flagged"]),   highlight_color="FFCCCC")
        if info.get("flagged_but_no_increase"):
            _write_sheet(wb, f"{crit_lbl}_FalsePos", _df_from_records(info["flagged_but_no_increase"]), highlight_color="FFF2CC")

    wb.save(str(xlsx_path))
    print(f"[xlsx_builder] XLSX written: {xlsx_path}")
