"""
modules/log_analyser.py
SAS log analysis module — mirrors %check_log in check_log.sas.

Scans a SAS log for ERROR and WARNING lines, deduplicates repeated messages,
classifies warnings as 'review' or 'known-benign', and can append two sheets
(SAS_Log_Errors, SAS_Log_Warnings) to an existing openpyxl workbook.
"""

import re
from pathlib import Path


# ── Known-benign warning patterns ────────────────────────────────────────────
_BENIGN_PATTERNS = [
    re.compile(p, re.IGNORECASE)
    for p in [
        r"variable \w+ (is uninitialized|not found)",
        r"multiple lengths were specified",
        r"numeric values have been converted",
        r"format \w+ was not found or could not be loaded",
        r"the data set .* has 0 observations",
        r"apparent symbolic reference .* not resolved",
    ]
]


def analyse_sas_log(log_path: str) -> dict:
    """
    Scan a SAS log for ERROR and WARNING lines.

    Rules (identical to %check_log in check_log.sas):
      - A line is an ERROR   if it matches /^\\s*ERROR:/i
      - A line is a WARNING  if it matches /^\\s*WARNING:/i

    Additional behaviour beyond the SAS macro:
      - Duplicate messages (same text on multiple lines) are deduplicated;
        extra line numbers are collected in 'also_at_lines'.
      - Warnings are classified as 'review' or 'known-benign'.

    Returns a dict suitable for JSON serialisation and report rendering.
    """
    error_pat   = re.compile(r"^\s*ERROR:",   re.IGNORECASE)
    warning_pat = re.compile(r"^\s*WARNING:", re.IGNORECASE)

    errors   = []
    warnings = []

    try:
        with open(log_path, encoding="utf-8", errors="replace") as f:
            for lineno, raw in enumerate(f, start=1):
                line = raw.rstrip("\n")
                if error_pat.match(line):
                    errors.append({"line_no": lineno, "text": line.strip()})
                elif warning_pat.match(line):
                    warnings.append({"line_no": lineno, "text": line.strip()})
        read_ok, read_error = True, ""
    except Exception as e:
        read_ok, read_error = False, str(e)

    def _dedup(items):
        seen, out = {}, []
        for item in items:
            t = item["text"]
            if t not in seen:
                seen[t] = item["line_no"]
                out.append(item)
            else:
                first = next(x for x in out if x["text"] == t)
                first.setdefault("also_at_lines", []).append(item["line_no"])
        return out

    errors_dedup   = _dedup(errors)
    warnings_dedup = _dedup(warnings)

    def _classify(text):
        return "known-benign" if any(p.search(text) for p in _BENIGN_PATTERNS) else "review"

    for w in warnings_dedup:
        w["category"] = _classify(w["text"])

    n_warn_review = sum(1 for w in warnings_dedup if w["category"] == "review")
    n_warn_benign = sum(1 for w in warnings_dedup if w["category"] == "known-benign")

    return {
        "log_path":          log_path,
        "read_ok":           read_ok,
        "read_error":        read_error,
        "n_errors":          len(errors),
        "n_warnings":        len(warnings),
        "n_errors_unique":   len(errors_dedup),
        "n_warnings_unique": len(warnings_dedup),
        "n_warnings_review": n_warn_review,
        "n_warnings_benign": n_warn_benign,
        "errors":            errors_dedup,
        "warnings":          warnings_dedup,
        "note": (
            r"Scan mirrors %check_log in check_log.sas: lines matching /^\s*ERROR:/i "
            r"or /^\s*WARNING:/i are captured. Identical messages are deduplicated; "
            "extra line numbers are listed in 'also_at_lines'. "
            "Warnings are classified as 'review' or 'known-benign'."
        ),
    }


def append_log_xlsx(xlsx_path: Path, log_path: str) -> None:
    """
    Append two sheets to an existing openpyxl workbook:
      SAS_Log_Errors   — all unique ERROR lines (red highlight)
      SAS_Log_Warnings — all unique WARNING lines (yellow/green by category)
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        print("  ⚠ openpyxl not available — cannot append log sheets to XLSX")
        return

    log_data = analyse_sas_log(log_path)

    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
    except Exception as e:
        print(f"  ⚠ Could not open XLSX to append log sheets: {e}")
        return

    header_fill  = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True)
    error_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    warn_rv_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    warn_bn_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def _also_str(item):
        also = item.get("also_at_lines", [])
        s    = ", ".join(str(x) for x in also[:10])
        if len(also) > 10:
            s += f" … (+{len(also)-10} more)"
        return s

    def _write(ws, headers, rows, fill_fn=None):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append(row)
            if fill_fn:
                f = fill_fn(row)
                if f:
                    for cell in ws[ws.max_row]:
                        cell.fill = f
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)

    # Errors sheet
    err_ws = wb.create_sheet("SAS_Log_Errors")
    err_rows = [[e["line_no"], _also_str(e), e["text"]] for e in log_data.get("errors", [])]
    if err_rows:
        _write(err_ws, ["Line_No", "Also_At_Lines", "ERROR_Text"], err_rows, lambda r: error_fill)
    else:
        _write(err_ws, ["Line_No", "Also_At_Lines", "ERROR_Text"],
               [["—", "—", "No ERRORs found in log ✅"]])
    print(f"  ✓ Sheet SAS_Log_Errors    ({len(err_rows)} unique ERROR(s))")

    # Warnings sheet
    warn_ws  = wb.create_sheet("SAS_Log_Warnings")
    warn_rows = [
        [w["line_no"], _also_str(w), w.get("category", "review"), w["text"]]
        for w in log_data.get("warnings", [])
    ]
    def _warn_fill(row):
        return warn_rv_fill if row[2] == "review" else warn_bn_fill
    if warn_rows:
        _write(warn_ws, ["Line_No", "Also_At_Lines", "Category", "WARNING_Text"], warn_rows, _warn_fill)
    else:
        _write(warn_ws, ["Line_No", "Also_At_Lines", "Category", "WARNING_Text"],
               [["—", "—", "—", "No WARNINGs found in log ✅"]])
    n_rv = log_data.get("n_warnings_review", 0)
    n_bn = log_data.get("n_warnings_benign", 0)
    print(f"  ✓ Sheet SAS_Log_Warnings  ({n_rv} to review, {n_bn} benign)")

    try:
        wb.save(str(xlsx_path))
        print(f"  ✓ XLSX updated: {xlsx_path}")
    except Exception as e:
        print(f"  ⚠ Could not save XLSX: {e}")
