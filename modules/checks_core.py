"""
modules/checks_core.py
Original QC check functions for the ADPC / ADNCA output dataset.

Each function accepts (df, mapping) or (df, mapping, sas_logic) and returns
a dict that is stored in qc_output_results.json under its section key.
"""

import re
from pathlib import Path

import pandas as pd

from .utils import col, col_list, is_blank, BAD_YEARS


# ─────────────────────────────────────────────────────────────────────────────
# SAS program parsing
# ─────────────────────────────────────────────────────────────────────────────

def parse_sas_flag_logic(sas_path: str, defaults: dict) -> dict:
    """
    Parse the production SAS program for whichc() and cmiss() calls that
    derive PKSUMXFL and NCAXFL / NCA01FL.

    defaults must contain 'pksumxfl_crits' and 'ncaxfl_ncaxrs' lists.
    """
    DEFAULT_PKSUMXFL_CRITS = defaults["pksumxfl_crits"]
    DEFAULT_NCAXFL_NCAXRS  = defaults["ncaxfl_ncaxrs"]

    result = {
        "pksumxfl_crits":    [],
        "ncaxfl_ncaxrs":     [],
        "nca01fl_logic":     "",
        "parse_warnings":    [],
        "parse_method":      "fallback_defaults",
        "raw_pksumxfl_stmt": "",
        "raw_ncaxfl_stmt":   "",
        "sas_path":          sas_path,
    }

    if not sas_path:
        result["parse_warnings"].append(
            "No SAS program provided. Default flag lists used — verify against your program."
        )
        result["pksumxfl_crits"] = DEFAULT_PKSUMXFL_CRITS[:]
        result["ncaxfl_ncaxrs"]  = DEFAULT_NCAXFL_NCAXRS[:]
        return result

    try:
        text = Path(sas_path).read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        result["parse_warnings"].append(f"Could not read SAS file: {e}. Default lists used.")
        result["pksumxfl_crits"] = DEFAULT_PKSUMXFL_CRITS[:]
        result["ncaxfl_ncaxrs"]  = DEFAULT_NCAXFL_NCAXRS[:]
        return result

    pksumxfl_match = re.search(
        r"whichc\s*\(\s*'Y'\s*,(.*?)\)\s*>\s*0\s+then\s+(?:do\s*;)?\s*PKSUMXFL\s*=\s*'Y'",
        text, re.IGNORECASE | re.DOTALL
    )
    if pksumxfl_match:
        raw_args = pksumxfl_match.group(1)
        result["raw_pksumxfl_stmt"] = pksumxfl_match.group(0).strip()
        crits = [c.upper() for c in re.findall(r"\bCRIT\d+FL\b", raw_args, re.IGNORECASE)]
        if crits:
            result["pksumxfl_crits"] = crits
            result["parse_method"]   = "whichc_cmiss"
        else:
            result["parse_warnings"].append(
                "whichc() found for PKSUMXFL but no CRITxxFL names extracted. Default list used."
            )
            result["pksumxfl_crits"] = DEFAULT_PKSUMXFL_CRITS[:]
    else:
        result["parse_warnings"].append(
            "whichc('Y', ...) > 0 then PKSUMXFL='Y' pattern not found. Default list used."
        )
        result["pksumxfl_crits"] = DEFAULT_PKSUMXFL_CRITS[:]

    ncaxfl_match = re.search(
        r"if\s+cmiss\s*\((.*?)\)\s*=\s*(\d+)\s+then\s+NCAXFL\s*=\s*'([^']*)'",
        text, re.IGNORECASE | re.DOTALL
    )
    if ncaxfl_match:
        raw_args = ncaxfl_match.group(1)
        result["raw_ncaxfl_stmt"] = ncaxfl_match.group(0).strip()
        ncaxrs = [v.upper() for v in re.findall(r"\bNCA\d+XRS\b", raw_args, re.IGNORECASE)]
        if ncaxrs:
            result["ncaxfl_ncaxrs"] = ncaxrs
            result["parse_method"]  = "whichc_cmiss"
        else:
            result["parse_warnings"].append(
                "cmiss() found for NCAXFL but no NCAxxXRS names extracted. Default list used."
            )
            result["ncaxfl_ncaxrs"] = DEFAULT_NCAXFL_NCAXRS[:]
    else:
        result["parse_warnings"].append(
            "cmiss(...) = N then NCAXFL pattern not found. Default list used."
        )
        result["ncaxfl_ncaxrs"] = DEFAULT_NCAXFL_NCAXRS[:]

    nca01_match = re.search(
        r"(if\s+RICHRFL\s*=\s*[\"']Y[\"']\s+AND\s+NCAXFL\s*=\s*[\"']\s*[\"']\s+then\s+NCA01FL\s*=\s*[\"']Y[\"'])",
        text, re.IGNORECASE
    )
    if nca01_match:
        result["nca01fl_logic"] = nca01_match.group(1).strip()
    else:
        result["nca01fl_logic"] = "RICHRFL = 'Y' AND NCAXFL = ' '  [pattern not found — assumed standard]"
        result["parse_warnings"].append("NCA01FL derivation pattern not found. Standard logic assumed.")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Specs comparison
# ─────────────────────────────────────────────────────────────────────────────

def check_specs_comparison(df: pd.DataFrame, specs_path: str) -> dict:
    """Compare ADNCA/ADPC specs sheet variable list against the dataset columns."""
    if not specs_path:
        return {"status": "No specifications file provided."}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(specs_path, read_only=True, data_only=True)
        su = {s.upper(): s for s in wb.sheetnames}
        specs_sheet = next((su[c] for c in ["ADNCA", "ADPC"] if c in su), None)
        if not specs_sheet:
            return {
                "status": (
                    f"Neither 'ADPC' nor 'ADNCA' sheet found in {Path(specs_path).name}. "
                    f"Sheets present: {', '.join(wb.sheetnames)}"
                )
            }
        specs_df = pd.read_excel(specs_path, sheet_name=specs_sheet, header=0)
        specs_df.columns = [str(c).strip() for c in specs_df.columns]
        cm = {c.upper(): c for c in specs_df.columns}
        name_col  = cm.get("NAME",  cm.get("VARIABLE"))
        label_col = cm.get("LABEL")
        core_col  = cm.get("CORE")
        if not name_col:
            return {"status": "Could not find a 'Name' or 'Variable' column in the specs sheet."}
        specs_df = specs_df.dropna(subset=[name_col])
        specs_df = specs_df[specs_df[name_col].astype(str).str.strip() != ""]
        specs_vars = [
            {
                "Name":  str(r[name_col]).strip().upper(),
                "Label": str(r[label_col]).strip() if label_col else "",
                "Core":  str(r[core_col]).strip()  if core_col  else "",
            }
            for _, r in specs_df.iterrows()
        ]
        ds = {c.upper() for c in df.columns}
        sn = {v["Name"] for v in specs_vars}
        return {
            "specs_sheet":               specs_sheet,
            "specs_path":                specs_path,
            "n_specs_vars":              len(specs_vars),
            "n_in_dataset":              sum(1 for v in specs_vars if v["Name"] in ds),
            "n_missing_from_dataset":    sum(1 for v in specs_vars if v["Name"] not in ds),
            "n_in_dataset_not_in_specs": len(ds - sn),
            "specs_vars":                specs_vars,
            "in_dataset":                [v for v in specs_vars if v["Name"] in ds],
            "missing_from_dataset":      [v for v in specs_vars if v["Name"] not in ds],
            "in_dataset_not_in_specs":   sorted(ds - sn),
        }
    except Exception as e:
        import traceback
        return {"status": f"Error reading specs file: {e}", "traceback": traceback.format_exc()}


# ─────────────────────────────────────────────────────────────────────────────
# Core QC checks
# ─────────────────────────────────────────────────────────────────────────────

def check_meta(df, data_path, mapping, user_override_misses):
    all_cols = list(df.columns)
    mapped     = {k: v.upper() for k, v in mapping.items() if v and v.upper() in df.columns}
    not_mapped = [k for k, v in mapping.items() if not v or v.upper() not in df.columns]
    return {
        "data_path":            str(data_path),
        "n_records":            int(len(df)),
        "n_columns":            int(len(all_cols)),
        "all_columns":          all_cols,
        "mapped_variables":     mapped,
        "not_mapped":           not_mapped,
        "user_override_misses": user_override_misses,
        "crit_fl_vars":  sorted([c for c in all_cols if re.match(r"^CRIT\d+FL$",  c, re.I)]),
        "crit_txt_vars": sorted([c for c in all_cols if re.match(r"^CRIT\d+$",    c, re.I)]),
        "nca_fl_vars":   sorted([c for c in all_cols if re.match(r"^NCA\d+FL$",   c, re.I)]),
        "nca_xrs_vars":  sorted([c for c in all_cols if re.match(r"^NCA\d+XRS$",  c, re.I)]),
    }


def check_subject_counts(df, mapping):
    uid  = col(df, "USUBJID", mapping)
    ph   = col(df, "PHASE",   mapping)
    arm  = col(df, "ACTARM",  mapping)
    test = col(df, "PCTEST",  mapping)
    def _grp(grp_col, label):
        if not grp_col or not uid: return []
        return [
            {label: str(k), "N_SUBJECTS": int(g[uid].nunique()), "N_RECORDS": int(len(g))}
            for k, g in df.groupby(grp_col)
        ]
    return {
        "total_subjects": int(df[uid].nunique()) if uid else None,
        "by_phase":       _grp(ph,   "PHASE"),
        "by_actarm":      _grp(arm,  "ACTARM"),
        "by_pctest":      _grp(test, "PCTEST"),
    }


def check_treatment_mapping(df, mapping):
    key_vars = ["TREAT", "SATRT", "ATRT", "AVISIT", "VISITCD",
                "PHASE", "ACTARM", "ACTARMCD", "TRT01A", "EXROUTE"]
    present = [v for v in key_vars if col(df, v, mapping)]
    if not present:
        return {"status": "No treatment mapping variables found in dataset.", "table": []}
    cols_actual = [col(df, v, mapping) for v in present]
    tbl = df[cols_actual].drop_duplicates().sort_values(cols_actual).reset_index(drop=True)
    tbl.columns = present
    return {
        "variables_found": present,
        "n_combinations":  int(len(tbl)),
        "table":           tbl.fillna("").astype(str).to_dict("records"),
    }


def check_route_table(df, mapping):
    rows = []
    for var in ["EXROUTE", "ECROUTE"]:
        c = col(df, var, mapping)
        if c:
            vc = df[c].value_counts(dropna=False).reset_index()
            vc.columns = ["VALUE", "N"]
            vc["VARIABLE"] = var
            rows.append(vc[["VARIABLE", "VALUE", "N"]])
    if not rows:
        return {"status": "EXROUTE and ECROUTE not found in dataset.", "table": []}
    return {"table": pd.concat(rows, ignore_index=True).fillna("").astype(str).to_dict("records")}


def check_crit_nca_crossref(df, mapping, sas_logic):
    pksumxfl_crits = set(sas_logic["pksumxfl_crits"])
    ncaxfl_ncaxrs  = set(sas_logic["ncaxfl_ncaxrs"])
    rows = []
    for crit_fl in sorted(df.columns):
        if not re.match(r"^CRIT\d+FL$", crit_fl, re.I): continue
        crit_base = re.sub(r"FL$", "", crit_fl, flags=re.I)
        crit_txt  = crit_base if crit_base in df.columns else ""
        meaning   = ""
        if crit_txt:
            vals = [str(v).strip() for v in df[crit_txt].dropna().unique() if str(v).strip()]
            meaning = vals[0] if vals else ""
        vc = df[crit_fl].value_counts(dropna=False)
        n_y, n_n = int(vc.get("Y", 0)), int(vc.get("N", 0))
        rows.append({
            "VARIABLE": crit_fl.upper(), "TYPE": "CRITxFL", "MEANING": meaning,
            "DRIVES_PKSUMXFL": "YES" if crit_fl.upper() in pksumxfl_crits else "no",
            "DRIVES_NCAXFL": "",
            "N_Y": n_y, "N_N": n_n, "N_MISSING": int(len(df)) - n_y - n_n,
        })
    for nca_xrs in sorted(df.columns):
        if not re.match(r"^NCA\d+XRS$", nca_xrs, re.I): continue
        n_pop = int((~is_blank(df[nca_xrs])).sum())
        rows.append({
            "VARIABLE": nca_xrs.upper(), "TYPE": "NCAxxXRS",
            "MEANING": str(df[nca_xrs].dropna().iloc[0]).strip() if n_pop > 0 else "",
            "DRIVES_PKSUMXFL": "",
            "DRIVES_NCAXFL": "YES" if nca_xrs.upper() in ncaxfl_ncaxrs else "no",
            "N_Y": n_pop, "N_N": "", "N_MISSING": int(len(df)) - n_pop,
        })
    for nca_fl in sorted(df.columns):
        if not re.match(r"^NCA\d+FL$", nca_fl, re.I): continue
        rows.append({
            "VARIABLE": nca_fl.upper(), "TYPE": "NCAxxFL", "MEANING": "",
            "DRIVES_PKSUMXFL": "", "DRIVES_NCAXFL": "",
            "N_Y": int(df[nca_fl].value_counts(dropna=False).get("Y", 0)),
            "N_N": "", "N_MISSING": "",
        })
    return {"crossref": rows}


def check_populations(df, mapping, sas_logic):
    uid    = col(df, "USUBJID",  mapping)
    nca01  = col(df, "NCA01FL",  mapping)
    pksumx = col(df, "PKSUMXFL", mapping)
    test   = col(df, "PCTEST",   mapping)
    visit  = col(df, "VISITCD",  mapping) or col(df, "AVISIT", mapping)
    pks_crits = sas_logic["pksumxfl_crits"]
    nca_ncaxrs = sas_logic["ncaxfl_ncaxrs"]
    result = {}

    if nca01:
        mask_nca = df[nca01].astype(str).str.strip() == "Y"
        nca_df   = df[mask_nca]
        crit_in_nca  = {c: int((nca_df[c].astype(str).str.strip() == "Y").sum())
                        for c in col_list(nca_df, pks_crits)
                        if (nca_df[c].astype(str).str.strip() == "Y").sum() > 0}
        ncaxrs_in_nca = {c: int((~is_blank(nca_df[c])).sum())
                         for c in col_list(nca_df, nca_ncaxrs)
                         if (~is_blank(nca_df[c])).sum() > 0}
        result["nca_population"] = {
            "n_records": int(mask_nca.sum()),
            "n_subjects": int(nca_df[uid].nunique()) if uid else None,
            "crit_y_within_nca": crit_in_nca,
            "ncaxrs_within_nca": ncaxrs_in_nca,
            "note": (
                "Records with NCA01FL='Y'. CRITxFL='Y' within this population "
                "are intentional inclusions — list and verify each."
            ),
        }
    else:
        result["nca_population"] = {"status": "NCA01FL not found in dataset", "n_records": 0}

    if pksumx:
        mask_sum = is_blank(df[pksumx])
        sum_df   = df[mask_sum]
        crit_in_sum = {c: int((sum_df[c].astype(str).str.strip() == "Y").sum())
                       for c in col_list(sum_df, pks_crits)
                       if (sum_df[c].astype(str).str.strip() == "Y").sum() > 0}
        result["summary_population"] = {
            "n_records": int(mask_sum.sum()),
            "n_subjects": int(sum_df[uid].nunique()) if uid else None,
            "crit_y_within_summary": crit_in_sum,
            "note": (
                "Records where PKSUMXFL is missing/empty. Any CRITxFL='Y' here "
                "means a flagged record contributed to summary statistics."
            ),
        }
    else:
        result["summary_population"] = {"status": "PKSUMXFL not found in dataset", "n_records": 0}

    if pksumx and nca01:
        mask_trough = is_blank(df[pksumx]) & (df[nca01].astype(str).str.strip() != "Y")
        by_visit = []
        if visit:
            for v, grp in df[mask_trough].groupby(visit):
                by_visit.append({
                    "VISITCD": str(v),
                    "N_RECORDS": int(len(grp)),
                    "N_SUBJECTS": int(grp[uid].nunique()) if uid else None,
                })
        result["trough_records"] = {
            "n_records": int(mask_trough.sum()),
            "by_visitcd": by_visit,
            "note": "PKSUMXFL missing AND NCA01FL != 'Y' — trough records in summary stats.",
        }
        mask_contra = (
            (df[pksumx].astype(str).str.strip() == "Y") &
            (df[nca01].astype(str).str.strip()  == "Y")
        )
        rows = []
        if mask_contra.sum() > 0 and uid and test and visit:
            rows = (df[mask_contra][[uid, test, visit] + col_list(df, [nca01, pksumx])]
                    .drop_duplicates().to_dict("records"))
        result["population_contradiction"] = {
            "n_records": int(mask_contra.sum()),
            "records":   rows,
            "note": "PKSUMXFL='Y' AND NCA01FL='Y' simultaneously — should never occur.",
        }
    return result


def check_dtype_flags(df, mapping, sas_path="", specs_vars=None):
    dtype_col = col(df, "DTYPE", mapping)
    dtype_in_sas, dtype_sas_note = None, ""
    if sas_path:
        try:
            lines = [ln.strip() for ln in Path(sas_path).read_text(encoding="utf-8", errors="replace").splitlines()
                     if "DTYPE" in ln.upper() and not ln.strip().startswith(("*", "/*"))]
            dtype_in_sas   = len(lines) > 0
            dtype_sas_note = (f"DTYPE referenced on {len(lines)} non-comment line(s) in SAS program."
                              if dtype_in_sas else "DTYPE not found in SAS program.")
        except Exception as e:
            dtype_sas_note = f"Could not read SAS file to check DTYPE: {e}"
    dtype_in_specs, dtype_specs_note = None, ""
    if specs_vars is not None:
        dtype_in_specs   = "DTYPE" in {v.upper() for v in specs_vars}
        dtype_specs_note = ("DTYPE is listed in the specifications."
                            if dtype_in_specs else
                            "DTYPE is NOT listed in the specifications.")
    if not dtype_col:
        return {"status": "DTYPE not found in dataset.",
                "dtype_in_sas": dtype_in_sas, "dtype_sas_note": dtype_sas_note,
                "dtype_in_specs": dtype_in_specs, "dtype_specs_note": dtype_specs_note}
    vc   = df[dtype_col].fillna("").astype(str).str.strip().value_counts(dropna=False)
    dist = [{"DTYPE": str(k), "N": int(v)} for k, v in vc.items()]
    nca01 = col(df, "NCA01FL", mapping)
    deleted_in_nca, deleted_in_nca_rows = 0, []
    if nca01:
        mask = (df[dtype_col].fillna("").astype(str).str.strip() == "DELETED") & \
               (df[nca01].astype(str).str.strip() == "Y")
        deleted_in_nca = int(mask.sum())
        if deleted_in_nca > 0:
            uid   = col(df, "USUBJID", mapping)
            test  = col(df, "PCTEST",  mapping)
            visit = col(df, "VISITCD", mapping)
            keep  = [c for c in [uid, test, visit, nca01, dtype_col] if c]
            deleted_in_nca_rows = df[mask][keep].to_dict("records")
    return {
        "distribution": dist,
        "deleted_in_nca_n": deleted_in_nca, "deleted_in_nca_rows": deleted_in_nca_rows,
        "dtype_in_sas": dtype_in_sas, "dtype_sas_note": dtype_sas_note,
        "dtype_in_specs": dtype_in_specs, "dtype_specs_note": dtype_specs_note,
        "note_deleted": "DELETED records must never have NCA01FL='Y'. Non-zero count is critical.",
    }


def check_mrrlt_duplicates(df, mapping):
    uid   = col(df, "USUBJID",  mapping)
    test  = col(df, "PCTEST",   mapping)
    visit = col(df, "VISITCD",  mapping) or col(df, "AVISIT", mapping)
    mrrlt = col(df, "MRRLT",    mapping)
    nrrlt = col(df, "NRRLT",    mapping)
    nca01 = col(df, "NCA01FL",  mapping)
    pksumx= col(df, "PKSUMXFL", mapping)
    aval  = col(df, "AVAL",     mapping)
    if not all([uid, test, visit, mrrlt]):
        return {"status": "Required variables not found (need USUBJID, PCTEST, VISITCD/AVISIT, MRRLT)."}

    def _find_dups(sub, label):
        key    = [uid, test, visit, mrrlt]
        counts = sub.groupby(key, dropna=False).size().reset_index(name="_cnt")
        dups   = counts[counts["_cnt"] > 1]
        if len(dups) == 0:
            return {"population": label, "n_duplicate_groups": 0, "n_duplicate_records": 0, "records": []}
        merged = dups.merge(sub, on=key, how="left")
        kc     = list(dict.fromkeys([c for c in key + ([nrrlt] if nrrlt else []) + ([aval] if aval else [])]))
        return {
            "population": label,
            "n_duplicate_groups":  int(len(dups)),
            "n_duplicate_records": int(len(merged)),
            "records": merged[kc + ["_cnt"]].sort_values(key).to_dict("records"),
        }

    result = {"all_records": _find_dups(df, "All records")}
    if nca01:
        result["nca_population"]     = _find_dups(df[df[nca01].astype(str).str.strip() == "Y"], "NCA01FL='Y'")
    if pksumx:
        result["summary_population"] = _find_dups(df[is_blank(df[pksumx])], "PKSUMXFL missing")
    return result


def check_erroneous_records(df, mapping):
    def _find(candidates):
        return next((c.upper() for c in candidates if c.upper() in df.columns), None)
    adtc_col = _find(["ADTC", "PCDTC"])
    arftdtc  = _find(["AREFDTC", "PCRFTDTC"])
    arftdtm  = _find(["ARFTDTM", "PCRFTDTM"])
    pcdtcefl = _find(["PCDTCEFL"])
    atm_col  = _find(["ATM"])
    uid      = col(df, "USUBJID", mapping)
    test     = col(df, "PCTEST",  mapping)
    visit    = col(df, "VISITCD", mapping) or col(df, "AVISIT", mapping)
    nrrlt    = col(df, "NRRLT",   mapping)
    keep_base = [c for c in [uid, test, visit, nrrlt] if c]
    result = {}
    for label, c in [("ADTC/PCDTC", adtc_col), ("AREFDTC/PCRFTDTC", arftdtc)]:
        if c:
            for yr in BAD_YEARS:
                mask = df[c].astype(str).str.contains(yr, na=False)
                result[f"bad_year_{yr}_{label.replace('/','_')}"] = {
                    "n": int(mask.sum()),
                    "records": df[mask][keep_base + [c]].to_dict("records") if mask.sum() > 0 else [],
                }
    if pcdtcefl:
        mask = df[pcdtcefl].astype(str).str.strip() == "Y"
        result["pcdtcefl_y"] = {
            "n": int(mask.sum()),
            "records": df[mask][keep_base + [pcdtcefl] + ([adtc_col] if adtc_col else [])].to_dict("records") if mask.sum() > 0 else [],
        }
    if arftdtm:
        mask = df[arftdtm].isna() | (df[arftdtm].astype(str).str.strip() == "")
        result["missing_arftdtm"] = {"n": int(mask.sum()), "note": "Missing dose datetime — MRRLT falls back to NRRLT"}
    if atm_col:
        try:
            atm_s    = pd.to_numeric(df[atm_col], errors="coerce")
            mask_num = atm_s.notna() & (atm_s == 0)
            mask_str = df[atm_col].astype(str).str.strip().isin(["0", "0.0", "00:00", "00:00:00"])
            mask     = mask_num | mask_str
        except Exception:
            mask = df[atm_col].astype(str).str.strip().isin(["0", "0.0", "00:00", "00:00:00"])
        result["atm_zero"] = {"n": int(mask.sum()), "note": "Sample time = 00:00 → NCA10XRS"}
    return result


def check_missing_dose(df, mapping, sas_logic):
    uid   = col(df, "USUBJID", mapping)
    test  = col(df, "PCTEST",  mapping)
    visit = col(df, "VISITCD", mapping) or col(df, "AVISIT", mapping)
    result = {}
    for flag, label in [("CRIT3FL", "Missed dose (CRIT3)"), ("CRIT7FL", "Incomplete dose (CRIT7)"),
                         ("MISSED_DOSE_FL", "MISSED_DOSE_FL"), ("RECORD_INCOMPLETE_DOSE", "RECORD_INCOMPLETE_DOSE")]:
        if flag.upper() in df.columns:
            mask = df[flag.upper()].astype(str).str.strip() == "Y"
            subjects = df[mask][uid].unique().tolist() if uid and mask.sum() > 0 else []
            result[flag] = {
                "label": label, "n_records": int(mask.sum()),
                "n_subjects": len(subjects), "subjects": [str(s) for s in subjects[:30]],
            }
    return result


def check_prepost_flags(df, mapping):
    uid = col(df, "USUBJID", mapping)
    result = {}
    for flag, label in [
        ("CRIT2FL",         "Post-dose taken pre-dose (CRIT2)"),
        ("CRIT10FL",        "Pre-dose taken post-dose (CRIT10)"),
        ("RECORD_PRE_POST", "Record: pre-dose nominally but collected post-dose"),
        ("RECORD_POST_PRE", "Record: post-dose nominally but negative ARRLT"),
    ]:
        if flag.upper() in df.columns:
            mask = df[flag.upper()].astype(str).str.strip() == "Y"
            subjects = df[mask][uid].unique().tolist() if uid and mask.sum() > 0 else []
            result[flag] = {
                "label": label, "n_records": int(mask.sum()),
                "n_subjects": len(subjects), "subjects": [str(s) for s in subjects[:30]],
            }
    return result


def check_vomiting(df, mapping):
    uid   = col(df, "USUBJID", mapping)
    visit = col(df, "VISITCD", mapping) or col(df, "AVISIT", mapping)
    result = {}
    for flag in ["CRIT12FL", "VOMFL", "PCVOMYN"]:
        if flag.upper() in df.columns:
            mask     = df[flag.upper()].astype(str).str.strip() == "Y"
            by_visit = []
            if visit and mask.sum() > 0:
                for v, grp in df[mask].groupby(visit):
                    by_visit.append({"VISITCD": str(v), "N_RECORDS": int(len(grp)),
                                     "N_SUBJECTS": int(grp[uid].nunique()) if uid else None})
            result[flag] = {
                "n_records":  int(mask.sum()),
                "n_subjects": int(df[mask][uid].nunique()) if uid and mask.sum() > 0 else 0,
                "by_visitcd": by_visit,
            }
    return result


def check_profile_types(df, mapping):
    pt   = "PROFILE_TYPE" if "PROFILE_TYPE" in df.columns else None
    ph   = col(df, "PHASE",   mapping)
    arm  = col(df, "ACTARM",  mapping)
    test = col(df, "PCTEST",  mapping)
    uid  = col(df, "USUBJID", mapping)
    if not pt:
        return {"status": "PROFILE_TYPE not found in dataset."}
    overall = df[pt].value_counts(dropna=False).reset_index()
    overall.columns = ["PROFILE_TYPE", "N"]
    by_group = []
    group_cols = [c for c in [ph, arm, test] if c]
    if group_cols and uid:
        grp = df.groupby(group_cols + [pt], dropna=False).agg(
            N_RECORDS=("USUBJID", "count"), N_SUBJECTS=(uid, "nunique")).reset_index()
        by_group = grp.fillna("").astype(str).to_dict("records")
    richrfl_dist = []
    if "RICHRFL" in df.columns:
        rv = df["RICHRFL"].fillna("").value_counts().reset_index()
        rv.columns = ["RICHRFL", "N"]
        richrfl_dist = rv.to_dict("records")
    return {
        "overall_distribution": overall.to_dict("records"),
        "by_group":             by_group,
        "richrfl_distribution": richrfl_dist,
    }


def check_richrfl_listrfl(df, mapping):
    nca01 = col(df, "NCA01FL", mapping)
    listr = "LISTRFL" if "LISTRFL" in df.columns else None
    richr = "RICHRFL" if "RICHRFL" in df.columns else None
    uid   = col(df, "USUBJID", mapping)
    result = {}
    if richr:
        result["richrfl_dist"] = {str(k): int(v) for k, v in df[richr].value_counts(dropna=False).items()}
    if listr:
        result["listrfl_dist"] = {str(k): int(v) for k, v in df[listr].value_counts(dropna=False).items()}
    if listr and nca01:
        mask = (df[listr].astype(str).str.strip() == "Y") & (df[nca01].astype(str).str.strip() == "Y")
        result["listrfl_and_nca01fl_y"] = {
            "n_records": int(mask.sum()),
            "note": "LISTRFL='Y' AND NCA01FL='Y' should never co-occur.",
            "records": df[mask][[uid] if uid else []].to_dict("records") if mask.sum() > 0 else [],
        }
    return result


def check_aprofile_alignment(df, mapping):
    uid   = col(df, "USUBJID",  mapping)
    aprof = col(df, "APROFILE", mapping)
    nca01 = col(df, "NCA01FL",  mapping)
    test  = col(df, "PCTEST",   mapping) or col(df, "PCTESTCD", mapping)
    visit = col(df, "AVISIT",   mapping) or col(df, "VISITCD",  mapping)
    if not aprof:
        return {"status": (
            "APROFILE not found in dataset. Profile grouping cannot be verified "
            "without reconstructing the key manually."
        )}
    if not nca01:
        return {"status": "NCA01FL not found — cannot check APROFILE alignment."}
    aprof_populated = ~is_blank(df[aprof])
    nca01_y         = df[nca01].astype(str).str.strip() == "Y"
    case1 = nca01_y & ~aprof_populated
    case2 = aprof_populated & ~nca01_y
    keep  = [c for c in [uid, test, visit, aprof, nca01] if c]
    def _sample(mask, n=50):
        return df[mask][keep].head(n).to_dict("records") if mask.sum() > 0 else []
    profile_summary = []
    if uid:
        pg = (df[aprof_populated].groupby(aprof)
              .agg(N_records=(uid, "count"),
                   N_nca01fl_y=(nca01, lambda x: (x.astype(str).str.strip() == "Y").sum()))
              .reset_index())
        pg.columns = ["APROFILE", "N_records", "N_NCA01FL_Y"]
        pg["All_NCA01FL_Y"] = pg["N_NCA01FL_Y"] == pg["N_records"]
        profile_summary = pg.head(200).to_dict("records")
    return {
        "aprofile_col":             aprof,
        "n_aprofile_populated":     int(aprof_populated.sum()),
        "n_aprofile_blank":         int((~aprof_populated).sum()),
        "n_nca01fl_y":              int(nca01_y.sum()),
        "n_nca01_y_aprofile_blank": int(case1.sum()),
        "case1_records":            _sample(case1),
        "case1_note": "NCA01FL='Y' but APROFILE blank — profile grouping is inconsistent.",
        "n_aprofile_nca01_not_y":   int(case2.sum()),
        "case2_records":            _sample(case2),
        "case2_note": "APROFILE populated but NCA01FL not 'Y' — review for validity.",
        "profile_summary": profile_summary,
    }
